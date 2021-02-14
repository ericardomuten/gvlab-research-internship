import pandas as pd
from math import *
from openpyxl import Workbook

#NEED TO BE CHANGE
subject = ["Edo1", "Edo2"]
num_of_subject = 2 #number of total subject
num_of_data = 40 #number of data for each cutted data excel
movement = ["notouch"]
num_of_movement = 1 #number of movement types
total_file = [5] #number of excel files for each corresponding movement
main_directory = "D:/TUAT/GV Lab Research Internship/Machine Learning/Data/Cut_Data/Sensor1/" #dont forget to change the sensor name!!!!

#STATIC VAR
name_of_coordinate = ["X", "Y", "Z"]

#DYNAMIC VAR
coordinate = 0 #0 for X, 1 for Y, 2 for Z
i = 0 #counting number of data in each file
counter = 0 #counting total file
move = 0 #counting number of movement types
total = 0 #counting the SUM of the data
total_square = 0 #counting the SUM of the (data)^2
all_sum = [0, 0, 0] #sum of total value (all of the movement types) for each axis
all_square_sum = [0, 0, 0] #sum of total (value)^2 (all of the movement types) for each axis
true_mean = [0, 0, 0]
true_stddev = [0, 0, 0]


#Setting up the temporary excel file & string file to save to notepad
wb = Workbook() #temporary excel file
ws = wb.active #choose the active worksheet
ws['A1'] = "Movement Type"
ws['B1'] = "Mean X"
ws['C1'] = "Std Dev X"
ws['D1'] = "Mean Y"
ws['E1'] = "Std Dev Y"
ws['F1'] = "Mean Z"
ws['G1'] = "Std Dev Z"
ws['H1'] = "Subject Name"
mean_column = ["B", "D", "F"]
stddev_column = ["C", "E", "G"]
out = ""

#Iteration to calculate Mean & Std Dev starts here
for people in range (0, num_of_subject): #subject change
    directory = main_directory + subject[people] + "/" + str(num_of_data)
    out += ("\n\nSubject : " + subject[people] + "\n")  # string to be printed
    ws['H' + str(1 + people*num_of_movement + 2)] = subject[people]
    while move < num_of_movement: #movement change
        out += ("\nMovement Type : " + movement[move] + "\n")
        ws['A' + str(1 + move + 2 + people*num_of_movement)] = movement[move]
        while coordinate < 3: #coordinate change
            while counter < total_file[move]: #file change
                file_read = pd.read_csv(directory + "/" + movement[move] + "_" + str(num_of_data) + "_" + str(counter) + ".csv")
                for i in range (0 , (num_of_data)): #data change
                    a = file_read[str(coordinate)][i]
                    total = total + a
                    all_sum[coordinate] = all_sum[coordinate] + a
                    total_square = total_square + (a*a)
                    all_square_sum[coordinate] = all_square_sum[coordinate] + (a*a)
                counter += 1
                i = 0
            mean = total/(num_of_data*total_file[move]) #expected value/mean of the data
            mean_square = total_square/(num_of_data*total_file[move]) #expected value/mean of the (data)^2
            stddev = sqrt(mean_square - (mean*mean)) #standard deviation (population)
            out += ("Mean " + name_of_coordinate[coordinate] + " : " + str(mean) + "\n")
            out += ("Std Dev " + name_of_coordinate[coordinate] + " : " + str(stddev) + "\n")
            ws[mean_column[coordinate] + str(1 + move + 2 + people*num_of_movement)] = mean
            ws[stddev_column[coordinate] + str(1 + move + 2 + people*num_of_movement)] = stddev

            i = 0  # counting number of data in each file
            counter = 0  # counting total file
            total = 0  # counting the SUM of the data
            total_square = 0  # counting the SUM of the (data)^2
            coordinate += 1
        coordinate = 0 #counting the coordinate
        move += 1
    move = 0 #counting the movement

#calculating total number of data for each axis (whole movement types)
N = 0
for j in range (0, num_of_movement):
    N = N + total_file[j]*num_of_data*num_of_subject
print (N)

#calculating whole mean & std dev
ws["A" + str(2)] = "all"
out += ("\nAll :\n")
for j in range (0, 3):
    true_mean[j] = all_sum[j]/N
    true_stddev[j] = sqrt((all_square_sum[j]/N) - (true_mean[j]*true_mean[j]))
    ws[mean_column[j] + str(2)] = true_mean[j]
    ws[stddev_column[j] + str(2)] = true_stddev[j]
    out += ("Mean " + name_of_coordinate[j] + " : " + str(true_mean[j]) + "\n")
    out += ("Std Dev " + name_of_coordinate[j] + " : " + str(true_stddev[j]) + "\n")

#writing the result to .txt file
write = main_directory + "Mean-Std_Dev_notouch" + "_" + str(num_of_data) + ".txt"
with open (write, "w") as file_write:
    file_write.write(out)

#saving the temporary excel file
wb.save(main_directory + "Mean-Std_Dev_notouch" + str(num_of_data) + ".xlsx")

print (true_mean)
print (true_stddev)
print ("Mean & Std Dev----end")